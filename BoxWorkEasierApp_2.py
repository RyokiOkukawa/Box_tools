import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import os
from tkinter import *
from tkinter import ttk, messagebox
from tkinter import filedialog
import threading
import pandas as pd
import re



excel_path = None  # エクセルのパスをグローバル変数として扱う
box_start_cnt = 0  # box作業の実行回数


# button1(参照)クリック時の処理
def button1_clicked():
    global excel_path
    fTyp = [("エクセルファイル", "*.xls?")]  # 選択できるファイルをエクセルのみに制限
    iDir = os.path.abspath(os.path.dirname(__file__))
    filepath = filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
    excel_path = filepath
    file1.set(filepath)


'''
button2(start)クリック時の処理
メインとは別スレッドで処理を実行
ここから
'''


def thread():
    """
    エクセルファイルが選択されているかをチェック
    選択されていなければファイル参照画面が出力
    """
    if excel_path == '' or excel_path is None:
        messagebox.showerror('エラー', 'エクセルファイルを選択してください。')
        button1_clicked()

        while file_flg:
            if excel_path != '' or excel_path is not None:
                file_flg = True
    '''
    box_programを開始する前の確認アラート
    '''
    number_of_executions = combo_box.get()  # 実行回数を取得
    ret = None
    if number_of_executions[0] == '1':
        ret = messagebox.askyesno('確認', '今回の実行回数は' + str(
            number_of_executions) + 'です。実行したら中断はできません。実行しますか？')

    if number_of_executions[0] == '3':
        ret = messagebox.askyesno('確認', '今回の実行回数は' + str(
            number_of_executions) + 'です。実行したら中断はできません。実行しますか？')

    if number_of_executions[0] == '5':
        ret = messagebox.askyesno('確認', '今回の実行回数は' + str(
            number_of_executions) + 'です。実行したら中断はできません。実行しますか？')

    if number_of_executions[0] == '8':
        ret = messagebox.askyesno('確認', '今回の実行回数は' + str(
            number_of_executions) + 'です。実行したら中断はできません。実行しますか？')

    if ret:
        global box_start_cnt

        root.title('BoxWorkEasierApp --実行中')  # 初期画面のタイトルを変更
        button2.config(state="disable")  # 初期画面のスタートボタンを使用不可状態に変更(多重処理をさせない)
        thread1 = threading.Thread(target=box_program)
        thread1.start()


def thread2():
    if excel_path == '' or excel_path is None:
        messagebox.showerror('エラー', 'エクセルファイルを選択してください。')
        button1_clicked()

        while file_flg:
            if excel_path != '' or excel_path is not None:
                file_flg = True

    ret = messagebox.askyesno('確認', '実行してもよろしいですか？')

    if ret:
        root.title('BoxWorkEasierApp --実行中')  # 初期画面のタイトルを変更
        button2.config(state="disable")  # 初期画面のスタートボタンを使用不可状態に変更(多重処理をさせない)
        button4.config(state="disable")
        thread2 = threading.Thread(target=box_format)
        thread2.start()


'''
エクセルファイルを整理するメソッド
'''


def box_program():
    """
    boxの作業経過を表示する画面
    """
    global box_start_cnt
    work_progress_display = Toplevel(root)  # 新しいウィンドウを作成
    work_progress_display.title(str(box_start_cnt) + '回目の実行')  # 実行回数をタイトルに表示
    work_progress_display.geometry("470x60")  # ウィンドウサイズ設定
    work_progress_display.resizable(False, False)  # 画面のリサイズを不可に設定

    console = Label(work_progress_display, width=40, anchor="w")  # 作業セルのコンソールフレーム作成
    console.place(x=4, y=5)  # フレームの座標
    console_par = Label(work_progress_display)  # 作業工数のコンソールフレーム
    console_par.place(x=300, y=5)  # フレームの座標
    progress = Label(work_progress_display, width=65, fg='#0000cd', bg='#ffffff', borderwidth=3, relief="ridge",
                     anchor="w")  # プログレスバーのフレーム作成
    progress.place(x=4, y=25)  # プログレスバーの座標

    '''
    ここからエクセルファイルをいじくるメイン処理
    '''
    book = openpyxl.load_workbook(excel_path)  # 任意のエクセルファイルをロード
    sheet = book.worksheets[0]  # シートの選択　デフォルトでは1枚目のシートを選択
    excel_name = excel_path.split('/')  # ブックの名前を取得
    match_reference_value_list = []  # 参照先行の要素リストを定義
    match_reference_cell_list = []  # 参照先行のセルリストを定義
    match_referrer_value_list = []  # 参照元行の要素リストを定義
    match_referrer_cell_list = []  # 参照元行のセルリストを定義
    progress_parameter = 38  # プログレスバーの文字数
    progress_bar = '■' * progress_parameter  # プログレスバーのテキスト
    p_count1 = 0  # プログレスバーの制御カウント(実行回数)
    p_count1_str = 0  # プログレスバーの制御カウント(出力文字数)
    referrer_row = select_sell.get()
    reference_row = select_sell2.get()

    '''
    エクセルシートからG列とN列が同じ値のセルを抽出するループ
    '''

    number_of_executions = combo_box.get()
    number_of_executions = int(number_of_executions[0])

    for i in range(number_of_executions):
        box_start_cnt += 1
        work_progress_display.title(str(box_start_cnt) + '回目の実行')
        elem_count = 0  # 上で作成した各種リストを取り出す用のカウント
        move_elems = 0  # 移動したコンテンツの数
        p_count2 = 0  # プログレスバーの文字数

        for row_referrer in sheet[referrer_row[0]]:  # 参照元行の要素を1つずつ取得
            console.config(text=row_referrer.coordinate)  # コンソールへの出力
            console_par.config(text='データ抽出中')  # コンソールへ出力
            progress.config(text=progress_bar, fg='#b0c4de')  # プログレスバーの出力
            '''
            10工数終わるたびにプログレスバーが黄色に点滅
            '''
            if p_count1 % 10 == 0:
                p_count1_str += 1
                progress.config(text=progress_bar, fg='#ffff00')

            row_referrer.value = str(row_referrer.value)  # None属性とint型を文字列とする(エラー回避)
            '''
            ここから取り出したG列の要素をN列と比較し同じ値のものはリストへappend
            '''
            row_referrer_replace = row_referrer.coordinate.replace(referrer_row[0], reference_row[0])

            for row_reference in sheet[reference_row[0]]:  # 参照先行の要素を1つずつ取得

                if row_reference.value is not None and row_referrer.value is not None:  # エラー回避のためNone属性を省く
                    row_reference.value = str(row_reference.value)  # None属性とint型を文字列とする(エラー回避)


                    if row_referrer.value in row_reference.value and sheet[
                        row_referrer_replace].value is not None:
                        if row_referrer.value.isdecimal():
                            print('これは数字です')
                            print(row_referrer.value)
                            if row_referrer.value == row_reference.value:
                                print('OK')
                                break
                            break
                        # G列の値がN列に含まれていたらリストへ格納
                        print(row_referrer.value)
                        match_reference_cell_list.append(row_reference.coordinate)  # N列のセル番号を格納
                        match_reference_value_list.append(row_reference.value)  # N列の値を格納
                        match_referrer_cell_list.append(row_referrer.coordinate)  # G列のセル番号を格納
                        match_referrer_value_list.append(row_referrer.value)  # G列の値を格納

                        p_count1 += 1  # 実行回数を+1

        '''
        ここから対象セルの置換処理
        '''

        # エクセルフォーマットが違う場合や該当するデータがなかった場合の例外処理
        try:
            progress_count = int(len(match_referrer_cell_list) / progress_parameter)  # プログレスバーの計算処理用変数定義1
            progress_count2 = int(len(match_referrer_cell_list) / progress_count)  # プログレスバーの計算処理用変数定義2
            progress_tmp_Value = progress_count  # プログレスバーの計算処理用変数定義3

        except ZeroDivisionError:
            messagebox.showerror('エラー', '参照元と同名の参照先の要素がありません。')
            work_progress_display.destroy()
            button2.config(state="normal")

        '''
        ここから
        セルの置換アルゴリズム
        '''
        test_list = []
        for copy_origin in match_referrer_cell_list:  # G列のセル番号リストを取得

            elem_copy = copy_origin.replace(referrer_row[0], reference_row[0])  # 例:G123をN123と列名を置き換えている
            elem_content = sheet[elem_copy].value  # 変数へN列の要素を代入

            # N列orG列がnullだった場合は置換ができないため除外
            if elem_content is not None and sheet[match_reference_cell_list[elem_count]].value is not None:
                # 置換処置 セルごとではなく、対象の文字列のみを置換するようにしている
                sheet[match_reference_cell_list[elem_count]] = (' ' + sheet[
                    match_reference_cell_list[elem_count]].value + ' ').replace(
                    ' ' + match_referrer_value_list[elem_count] + ' ',
                    ' ' + elem_content + ' ').strip()
                test_list.append(sheet[match_reference_cell_list[elem_count]].value)

                # コンソール画面へ置換した値を出力 (特に意味ないけどそれっぽいからやった)
                console.config(text=sheet[match_reference_cell_list[elem_count]].value)
                console_par.config(
                    text=str(elem_count) + '/' + str(len(match_referrer_cell_list)))  # 処理済みのファイル数/処理対象の全ファイルをコンソールへ表示
                move_elems += 1  # プログレスバーの文字数を+1

            # メンバーの重複を削除
            dupli_delete = sheet[match_reference_cell_list[elem_count]].value
            dupli_delete = dupli_delete.split()
            dupli_add = set(dupli_delete)
            values = ''
            for value in dupli_add:
                values = values + value + ' '
            sheet[match_reference_cell_list[elem_count]] = values
            '''
            プログレスバー専用の処理
            '''
            if elem_count == progress_count2:
                p_count2 += 1
                progress_count2 += progress_tmp_Value
            progress.config(text='■' * p_count2)
            elem_count += 1

        if bln.get():
            messagebox.showinfo('結果', '検出した項目数:' + str(p_count1) + '\n総移動数:' + str(move_elems))

        if i == number_of_executions - 1:
            work_progress_display.destroy()
            button2.config(state="normal")  # スタートボタンが押せる状態になる

        try:
            book.save(excel_name[-1])  # 変更されたエクセルファイルを保存(直下に保存されるため別ファイルが生成される)

        except PermissionError:
            messagebox.showerror('エラー', '編集しているエクセルファイルが開かれている可能性があります。')
        i += 1
        '''
        メインの処理完了
        '''

    try:
        book.save(excel_name[-1])  # 変更されたエクセルファイルを保存(直下に保存されるため別ファイルが生成される)

    except PermissionError:
        messagebox.showerror('エラー', '編集しているエクセルファイルが開かれている可能性があります。')
        work_progress_display.destroy()
        button2.config(state="normal")

    root.title('Box作業ツール')  # GUIアプリのタイトルを初期値に戻す

    '''
    処理が終了した際に表示する結果画面
    '''
    if bln.get() is False:
        exit_flg = messagebox.showinfo(str(box_start_cnt) + '回目の実行結果',
                                       '検査した項目数:' + str(elem_count) + '\n総移動数:' + str(move_elems))
        if exit_flg == 'ok':  # okを押下後コンソールも閉じる
            work_progress_display.destroy()
        button2.config(state="normal")  # スタートボタンが押せる状態になる

    '''
    button2(start)クリック時の処理
    ここまで
    '''

def box_format():
    format_display = Toplevel(root)  # 新しいウィンドウを作成
    format_display.title('データをフォーマット中')  # 実行回数をタイトルに表示
    format_display.geometry("150x40")  # ウィンドウサイズ設定
    format_display.resizable(False, False)  # 画面のリサイズを不可に設定

    console = Label(format_display, width=40, anchor="w")  # 作業セルのコンソールフレーム作成
    console.place(x=4, y=5)
    book = openpyxl.load_workbook(excel_path)  # 任意のエクセルファイルをロード
    origin_data = book.worksheets[0]
    format_data = book.create_sheet(title='グループメンバー一覧')
    fill = PatternFill(patternType='solid',
                       fgColor='92D050', bgColor='92D050')
    dust_id_list_alnum = []
    dust_group_list_alnum = []
    dust_group_list_jp = []
    dust_id_list_jp = []

    alnum = re.compile(r'^[a-zA-Z0-9_!-/:-@-`{-~]+$')


    for value in range(1, 20000):
        format_data.row_dimensions[value].height = 18.75

    group_row = select_sell.get()
    member_row = select_sell2.get()

    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    format_data.column_dimensions['A'].width = 9.63
    format_data.column_dimensions['B'].width = 32.25
    format_data.column_dimensions['C'].width = 29.25
    format_data.column_dimensions['D'].width = 12.63
    format_data.column_dimensions['E'].width = 14.63
    format_data.column_dimensions['F'].width = 40.75
    format_data.column_dimensions['G'].width = 13.38
    format_data.column_dimensions['H'].width = 5
    format_data.column_dimensions['I'].width = 36.5

    format_data['A1'] = '◆Boxグループメンバーリスト'
    format_data['A3'] = 'No.'
    format_data['A3'].fill = fill
    format_data['B3'] = 'グループ名'
    format_data['B3'].fill = fill
    format_data['C3'] = 'メンバー社員番号'
    format_data['C3'].fill = fill
    format_data['D3'] = 'グループID'
    format_data['D3'].fill = fill
    format_data['E3'] = 'Boxユーザー名'
    format_data['E3'].fill = fill
    format_data['F3'] = 'メールアドレス'
    format_data['F3'].fill = fill
    format_data['G3'] = 'ユーザーID'
    format_data['G3'].fill = fill
    format_data['H3'] = '結果'
    format_data['H3'].fill = fill
    format_data['I3'] = '備考'
    format_data['I3'].fill = fill
    i = 4
    for group_name in origin_data[group_row[0]]:
        user_num_cell = group_name.coordinate.replace(group_row[0], member_row[0])
        user_num = origin_data[user_num_cell].value
        format_data['A' + str(i)] = i - 3
        if user_num is not None:
            # print(user_num)
            user_num_list = user_num.split()
            for user_num in range(len(user_num_list)):
                # print(user_num)
                # print(len(user_num_list))
                if user_num_list[user_num].isdecimal():
                    format_data['B' + str(i)] = '5435_' + str(group_name.value)
                    format_data['B' + str(i)].fill = PatternFill(patternType='solid'
                                                                 , fgColor='EAEDC9', bgColor='EAEDC9')
                    console.config(text=group_name.value)
                    format_data['C' + str(i)] = user_num_list[user_num]
                    format_data['C' + str(i)].fill = PatternFill(patternType='solid'
                                                                 , fgColor='EAEDC9', bgColor='EAEDC9')
                    print(str(i))
                    i += 1
                elif alnum.search(user_num_list[user_num]):
                    dust_group_list_alnum.append(group_name.value)
                    dust_id_list_alnum.append(user_num_list[user_num])

                else:
                    dust_group_list_jp.append(group_name.value)
                    dust_id_list_jp.append(user_num_list[user_num])

    print(dust_id_list_jp)
    for row in format_data:
        for cell in row:
            if format_data[cell.coordinate].value:
                format_data[cell.coordinate].border = border

    exit_flg = messagebox.showinfo('結果', '書き出した行数' + str(i - 4))
    if exit_flg == 'ok':  # okを押下後コンソールも閉じる
        out_list = [[0 for j in range(2)] for s in range(len(dust_id_list_alnum))]  # 多次元配列を作成
        for i in range(len(out_list)):  # 多次元配列に要素を詰め込んでいくここはkey値を指定
            for j in range(len(out_list[i])):  # value値を格納
                if j == 0:  # 各要素の住所に入れる条件分岐
                    out_list[i][j] = dust_group_list_alnum[i]  # パスの住所は0
                if j == 1:
                    out_list[i][j] = dust_id_list_alnum[i]  # エラーパスの住所は1

        out_list2 = [[0 for j in range(2)] for s in range(len(dust_id_list_jp))]  # 多次元配列を作成
        for i in range(len(out_list2)):  # 多次元配列に要素を詰め込んでいくここはkey値を指定
            for j in range(len(out_list2[i])):  # value値を格納
                if j == 0:  # 各要素の住所に入れる条件分岐
                    out_list2[i][j] = dust_group_list_jp[i]  # パスの住所は0
                if j == 1:
                    out_list2[i][j] = dust_id_list_jp[i]  # エラーパスの住所は1

        df = pd.DataFrame(out_list, columns=['グループ名', 'ユーザID'])
        df.to_excel('summary_alnum.xlsx')
        df = pd.DataFrame(out_list2, columns=['グループ名', 'ユーザID'])
        df.to_excel('summary_kana.xlsx')
        book.save('グループ設定一覧.xlsx')
        format_display.destroy()
        button4.config(state="normal")  # スタートボタンが押せる状態になる
        button2.config(state="normal")
        root.title('Box作業ツール')  # GUIアプリのタイトルを初期値に戻す


def quit_app():  # button3クリック時の処理
    root.destroy()


if __name__ == '__main__':
    # rootの作成
    root = Tk()
    root.title('BoxWorkEasierApp')
    root.resizable(False, False)

    # Frame1の作成
    frame1 = ttk.Frame(root, padding=10)
    frame1.grid()

    # 参照ボタンの作成
    button1 = ttk.Button(root, text=u'参照', command=button1_clicked)
    button1.grid(row=0, column=3)

    # ラベルの作成
    # 「ファイル」ラベルの作成
    s = StringVar()
    s.set('ファイル>>')
    label1 = ttk.Label(frame1, textvariable=s)
    label1.grid(row=0, column=0)

    # 参照ファイルパス表示ラベルの作成
    file1 = StringVar()
    file1_entry = ttk.Entry(frame1, textvariable=file1, width=50)
    file1_entry.grid(row=0, column=2)

    # フレーム2作成
    frame2 = ttk.Frame(root, padding=(0, 5))
    frame2.grid(row=1)

    # 対象セルを選択(コピー対象の名前)
    select_sell_str = ttk.Label(frame2)
    select_sell_str.config(text='グループの列')
    select_sell_str.pack(side=LEFT)

    # 対象セルを選択（ドロップダウンリスト)
    select_sell = ttk.Combobox(frame2, width="4")
    select_sell["value"] = \
        ("A列", "B列", "C列", "D列", "E列",
         "F列", "G列", "H列", "I列", "J列",
         "K列", "L列", "M列", "N列", "O列",
         "P列", "Q列", "R列", "S列", "T列",
         "U列", "V列", "W列", "X列", "Y列", "Z列")

    select_sell.current(0)
    select_sell.pack(side=LEFT)

    # 対象セルを選択(コピー対象の名前)
    select_sell_str2 = ttk.Label(frame2)
    select_sell_str2.config(text='メンバーの列')
    select_sell_str2.pack(side=LEFT)

    # 対象セルを選択（ドロップダウンリスト)
    select_sell2 = ttk.Combobox(frame2, width="4")
    select_sell2["value"] = \
        ("A列", "B列", "C列", "D列", "E列",
         "F列", "G列", "H列", "I列", "J列",
         "K列", "L列", "M列", "N列", "O列",
         "P列", "Q列", "R列", "S列", "T列",
         "U列", "V列", "W列", "X列", "Y列", "Z列")

    select_sell2.current(0)
    select_sell2.pack(side=LEFT)

    null_label = ttk.Label(frame2)
    null_label.config(text='   ')
    null_label.pack(side=LEFT)

    bln = BooleanVar()
    bln.set(True)
    result_every_time = ttk.Checkbutton(frame2, variable=bln, text='結果を実行毎に出力する')
    result_every_time.pack(side=LEFT)

    # Frame3の作成
    frame3 = ttk.Frame(root, padding=(0, 5))
    frame3.grid(row=2)

    button4 = ttk.Button(frame3, text='フォーマット', command=thread2)
    button4.pack(side=RIGHT)

    # Cancelボタンの作成
    button3 = ttk.Button(frame3, text='Cancel', command=quit_app)
    button3.pack(side=RIGHT)

    # Startボタンの作成
    button2 = ttk.Button(frame3, text='Start', command=thread)
    button2.pack(side=RIGHT)

    # 実行回数テキスト
    label2 = ttk.Label(frame3)
    label2.config(text='実行回数:')
    label2.pack(side=LEFT)

    # 実行回数を選択するドロップダウンリスト
    combo_box = ttk.Combobox(frame3, width="3")
    combo_box["value"] = ("1回", "3回", "5回", "8回")
    combo_box.current(0)
    combo_box.pack(side=RIGHT)

    root.mainloop()

'''
作成

BI部第一インテグレーション課
奥川凌貴

最終更新日:2020/08/14
'''
